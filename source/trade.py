"""
Created on Jan 15, 2017
◘
@author: Souvik
"""

import os
import openpyxl as xl
from datetime import datetime
from shutil import copy2
import timelog
import numpy as np
from opthist import OptionsHist

class TradeTest:
    """Main Class for performing a trading backtest"""

    def __init__(self, root, dataFile, optionsDB):

        # variables

        self.nth_expiry, self.nth_strike, self.midmonth_cutoff, self.entrytime = 0, 0, 10, 'sod'

        self.dataFile = dataFile
        self.testFile = self.dataFile.replace(".xlsx", "_Test_" + str(datetime.now().strftime('%Y-%m-%d %H.%M.%S')) +
                                              ".xlsx")
        self.optionsFile = optionsDB

        # init

        # starttime = timelog.now()

        os.chdir(root)
        copy2(self.dataFile, self.testFile) # Create test file for analysis

        print('TradeTest.__init__', self.testFile)

        self.dataHist = DataHist(self.testFile)

        self.oh = OptionsHist(optionsDB)


    def signals(self, latitude=0):
        """ Return trading signals
        Sample Return Format: ["Buy", None, None, "Sell", None, ....]
        Call Format: signals(percentage latitude to compare close against lower high an higher low)
        """

        buy = np.logical_and(self.dataHist.value['BarUpDown'] == 1,
                              self.dataHist.value['Close'] > self.dataHist.value['LH Value'] * (1 + latitude))
        sell = np.logical_and(self.dataHist.value['BarUpDown'] == -1,
                              self.dataHist.value['Close'] < self.dataHist.value['HL Value'] * (1 - latitude))

        signals = []
        bought, sold = False, False

        for i in range(0, self.dataHist.barcount):
            if buy[i]:
                if not bought:
                    signals.append("Buy")
                    bought, sold = True, False
                else:
                    signals.append(None)
            elif sell[i]:
                if not sold:
                    signals.append("Sell")
                    sold, bought = True, False
                else:
                    signals.append(None)
            else:
                signals.append(None)

        return signals


    def trades(self, signals):
        """ Return trades
        Return Format: [{'Symbol' : <>, 'Date' : <>, 'Close' : <>, 'Signal' : <>}, {...}, {...}, ....]
        Call Format: trades(returned list from signals() method)
        """

        trades = []

        for i in range(0,self.dataHist.barcount):
            if signals[i] in ("Buy", "Sell"):
                trades.append(
                    {
                        'Symbol': self.dataHist.value['Symbol'][i],
                        'Date': self.dataHist.value['Date'][i],
                        'Close': self.dataHist.value['Close'][i],
                        'Signal': signals[i]
                    }
                )

        return trades


    def sel_options(self, trades):

        optTrades = []

        for i in range(0, len(trades)):

            optSymbol = 'NIFTY' if trades[i]['Symbol'] == 'NSENIFTY' else trades[i]['Symbol']
            optType = "CE" if trades[i]['Signal'] == "Buy" else "PE"

            optTrades.append(
                self.oh.immediate_opt(
                    r"'{}'".format(optSymbol),
                    type=r"'{}'".format(optType),
                    date=r"'{}'".format(trades[i]['Date']),
                    close=trades[i]['Close']
                )
            )


    def signal_data(self, bar, signal):
        """ Returns signal data (similar to trades function
        Return Format: {'Symbol' : <>, 'Date' : <>, 'Close' : <>, 'Signal' : <>}
        Call Format: signal_data(bar_index, signal)
        """

        return {
            'Symbol': self.dataHist.value['Symbol'][bar],
            'Date': self.dataHist.value['Date'][bar],
            'Close': self.dataHist.value['Close'][bar],
            'Signal': signal
        }


    def sel_option(self, signal_data):
        """Return selected option based on signal_data"""

        optSymbol = 'NIFTY' if signal_data['Symbol'] == 'NSENIFTY' else signal_data['Symbol']
        optType = "CE" if signal_data['Signal'] == "CE Buy" else "PE"

        optiondata = self.oh.nth_opt(self.nth_expiry, self.nth_strike, self.midmonth_cutoff, self.entrytime,
            symbol=r"'{}'".format(optSymbol),
            type=r"'{}'".format(optType),
            date=r"'{}'".format(signal_data['Date']),
            close=signal_data['Close']
        )

        return optiondata


    def option_exit_data(self, i, option_entry):
        """Return option data on exit date based on option_entry"""

        exit_date = self.dataHist.bar(i)['Date']
        exit_data = self.oh.option_exit_data(option_entry, exit_date)

        return exit_data


    def check_exit_target(self, entry_price, date, exit_target, option_data):

        target_exit_price = entry_price * exit_target if exit_target > 0 else None
        target_exit = None

        if date in option_data['date'] and target_exit_price is not None:
            if target_exit_price <= option_data['open'][option_data['date'].index(date)]:
                target_exit = 'Target - Open'
            elif target_exit_price <= option_data['high'][option_data['date'].index(date)]:
                target_exit = 'Target - High'

        return target_exit





    def option_data(self, exit_date, option_data):

        return_data = {}

        bar_count = option_data['date'].index(exit_date) + 1
        for column, data in option_data.items():
            return_data[column] = option_data[column][:bar_count]

        return return_data


    def write_trade_data(self, trade):

        entry_date = None

        for signal, data in trade.items():
            if signal == 'entry':
                entry_data = {'Entry Type': data['type'], 'Expiry': data['expiry'], 'Strike Price': data['strikeprice'],
                              'Entry Date': data['date'], 'Entry Open': data['open'], 'Entry High': data['high'],
                              'Entry Low': data['low'], 'Entry Close': data['close'], 'Entry LTP': data['ltp']
                              }
                self.dataHist.write_data_by_date(data['date'], entry_data)
                entry_date = data['date']
            elif signal == 'exit':
                entry_data = {'Exit Date': data['date']}
                self.dataHist.write_data_by_date(entry_date, entry_data) # Update exit date in entry bar
                exit_data = {'Prev_Trade Expiry': data['expiry'], 'Prev_Trade Strike Price': data['strikeprice'],
                             'Prev_Trade Open': data['open'], 'Prev_Trade High': data['high'],
                             'Prev_Trade Low': data['low'], 'Prev_Trade Close': data['close'],
                             'Prev_Trade LTP': data['ltp'], 'Exit Type': data['exittype']
                             }
                self.dataHist.write_data_by_date(data['date'], exit_data)

    def write_trade_data_all(self, trade):

        entry_date = None

        for datatype, data in trade.items():
            if datatype == 'entry':
                entry_data = {'Entry Type': data['type'], 'Expiry': data['expiry'],
                              'Strike Price': data['strikeprice'],
                              'Entry Date': data['date'], 'Entry Open': data['open'], 'Entry High': data['high'],
                              'Entry Low': data['low'], 'Entry Close': data['close'], 'Entry LTP': data['ltp']
                              }
                self.dataHist.write_data_by_date(data['date'], entry_data)
                entry_date = data['date']
            elif datatype == 'exit':
                entry_data = {'Exit Date': data['date']}
                self.dataHist.write_data_by_date(entry_date, entry_data)  # Update exit date in entry bar
                exit_data = {'Exit Type': data['exittype']}
                self.dataHist.write_data_by_date(data['date'], exit_data)
            elif datatype == 'option_data':
                for i in range(0, len(data['date'])):
                    daily_data = {'Prev_Trade Expiry': data['expiry'][i],
                                  'Prev_Trade Strike Price': data['strikeprice'][i],
                                  'Prev_Trade Open': data['open'][i], 'Prev_Trade High': data['high'][i],
                                  'Prev_Trade Low': data['low'][i], 'Prev_Trade Close': data['close'][i],
                                  'Prev_Trade LTP': data['ltp'][i]
                                 }
                    self.dataHist.write_data_by_date(data['date'][i], daily_data)


    def backtest(self, start=0, end=0, exit_target=0, nth_expiry=0, nth_strike=0, midmonth_cutoff=10, entrytime='sod'):
        """Return historical trades
        Return format: {entry date in YYYY-MM-DD: {"entry": {option data from sel_option()},
                                                   "exit": {option data from option_exit_data()}
                                 },
                        entry date in YYYY-MM-DD: {"entry": {option data from sel_option()},
                                                   "exit": {option data from option_exit_data()}
                                 },......
                        }
        Call format: backtest(start=0,              start date for backtest in YYYY-MM-DD
                              end=0,                end date for backtest in YYYY-MM-DD
                              exittarget=2,         exit target in multiple of entry price
                              nth_expiry=0,         option expiry month to be selected
                              nth_strike=0,         option strike price to be selected
                              midmonth_cutoff=10    cut-off date to be used to select immediate expiry or next
                              entrytime='sod'       entry on same end of day ('eod') or next day start of day ('sod')
                            )
        """

        self.dataHist.add_column_names('Entry Type', 'Expiry', 'Strike Price', 'Entry Date', 'Entry Open', 'Entry High',
                                       'Entry Low', 'Entry Close', 'Entry LTP', 'Exit Date', 'Prev_Trade Expiry',
                                       'Prev_Trade Strike Price', 'Prev_Trade Open', 'Prev_Trade High', 'Prev_Trade Low',
                                       'Prev_Trade Close', 'Prev_Trade LTP', 'Exit Type')

        if start == 0 or start < self.dataHist.value['Date'][0]:
            start = self.dataHist.value['Date'][0]
        if end == 0 or end > self.dataHist.value['Date'][self.dataHist.barcount - 1]:
            end = self.dataHist.value['Date'][self.dataHist.barcount - 1]

        print('Running backtest from', start, 'to', end, '. . . .')

        self.nth_expiry, self.nth_strike, self.midmonth_cutoff, self.entrytime = \
            nth_expiry, nth_strike, midmonth_cutoff, entrytime

        buy = np.logical_and(self.dataHist.value['BarUpDown'] == 1,
                              self.dataHist.value['Close'] > self.dataHist.value['LH Value'] * (1 + latitude))
        sell = np.logical_and(self.dataHist.value['BarUpDown'] == -1,
                              self.dataHist.value['Close'] < self.dataHist.value['HL Value'] * (1 - latitude))

        trades = {}
        CE_bought, CE_sold, PE_bought, PE_sold = False, False, False, False
        CE_entry_date, PE_entry_date = None, None
        CE_option_data, PE_option_data = None, None
        on_expiry, on_target_price, next_trade_bar = None, None, None

        for i in range(self.dataHist.barindex[start], self.dataHist.barindex[end]):

            if CE_bought:
                # print('TradeTest.backtest', trades[CE_entry_date])
                on_expiry = self.dataHist.value['Date'][i] == CE_option_data['date'][len(CE_option_data['date'])-1]
                on_target_price = self.check_exit_target(trades[CE_entry_date]['entry']['open'],
                                                         self.dataHist.value['Date'][i], exit_target,
                                                         CE_option_data)
                if on_expiry or on_target_price is not None:
                    if on_expiry:
                        # Exit CE option on expiry date
                        trades[CE_entry_date]['exit'] = {'date': self.dataHist.value['Date'][i], 'exittype': 'Expiry'}
                        next_trade_bar = i - 1
                    elif on_target_price is not None:
                        # Exit CE option if target price met
                        trades[CE_entry_date]['exit'] = {'date': self.dataHist.value['Date'][i],
                                                         'exittype': on_target_price}
                        next_trade_bar = i if on_target_price == 'Target - High' else i - 1
                    trades[CE_entry_date]['option_data'] = self.option_data(trades[CE_entry_date]['exit']['date'],
                                                                            CE_option_data)
                    # CE_bought, CE_sold = False, True
                    self.write_trade_data_all(trades[CE_entry_date])
                    # Enter next CE option on expiry date
                    selected_option = self.sel_option(self.signal_data(next_trade_bar, "CE Buy"))  # Entry CE Option
                    CE_entry_date = selected_option['date']
                    trades[CE_entry_date] = {}
                    trades[CE_entry_date]['entry'] = selected_option
                    CE_option_data = self.oh.option_data_post_entry(trades[CE_entry_date])
                    CE_bought, CE_sold = True, False
            elif PE_bought:
                # print('TradeTest.backtest', trades[PE_entry_date])
                on_expiry = self.dataHist.value['Date'][i] == PE_option_data['date'][len(PE_option_data['date']) - 1]
                on_target_price = self.check_exit_target(trades[PE_entry_date]['entry']['open'],
                                                         self.dataHist.value['Date'][i], exit_target,
                                                         PE_option_data)
                if on_expiry or on_target_price is not None:
                    if on_expiry:
                        # Exit PE option on expiry date
                        trades[PE_entry_date]['exit'] = {'date': self.dataHist.value['Date'][i], 'exittype': 'Expiry'}
                        next_trade_bar = i - 1
                    elif on_target_price is not None:
                        # Exit PE option if target price met
                        trades[PE_entry_date]['exit'] = {'date': self.dataHist.value['Date'][i],
                                                         'exittype': on_target_price}
                        next_trade_bar = i if on_target_price == 'Target - High' else i - 1
                    trades[PE_entry_date]['option_data'] = self.option_data(trades[PE_entry_date]['exit']['date'],
                                                                            PE_option_data)
                    # PE_bought, PE_sold = False, True
                    self.write_trade_data_all(trades[PE_entry_date])
                    # Enter next PE option on expiry date
                    selected_option = self.sel_option(self.signal_data(next_trade_bar, "PE Buy"))  # Entry PE Option
                    PE_entry_date = selected_option['date']
                    trades[PE_entry_date] = {}
                    trades[PE_entry_date]['entry'] = selected_option
                    PE_option_data = self.oh.option_data_post_entry(trades[PE_entry_date])
                    PE_bought, PE_sold = True, False



            if buy[i]:
                if PE_bought:
                    trades[PE_entry_date]['exit'] = {'date': self.dataHist.value['Date'][i+1], 'exittype': 'Switch'}
                    PE_bought, PE_sold = False, True
                    trades[PE_entry_date]['option_data'] = self.option_data(trades[PE_entry_date]['exit']['date'],
                                                                            PE_option_data)
                    self.write_trade_data_all(trades[PE_entry_date])
                if not CE_bought:
                    selected_option = self.sel_option(self.signal_data(i, "CE Buy")) # Entry CE Option
                    CE_entry_date = selected_option['date']
                    trades[CE_entry_date] = {}
                    trades[CE_entry_date]['entry'] = selected_option
                    CE_bought, CE_sold = True, False
                    CE_option_data = self.oh.option_data_post_entry(trades[CE_entry_date])
            elif sell[i]:
                if CE_bought:
                    trades[CE_entry_date]['exit'] = {'date': self.dataHist.value['Date'][i+1], 'exittype' : 'Switch'}
                    CE_bought, CE_sold = False, True
                    trades[CE_entry_date]['option_data'] = self.option_data(trades[CE_entry_date]['exit']['date'],
                                                                            CE_option_data)
                    self.write_trade_data_all(trades[CE_entry_date])
                if not PE_bought:
                    selected_option = self.sel_option(self.signal_data(i, "PE Buy")) # Entry PE Option
                    PE_entry_date = selected_option['date']
                    trades[PE_entry_date] = {}
                    trades[PE_entry_date]['entry'] = selected_option
                    PE_bought, PE_sold = True, False
                    PE_option_data = self.oh.option_data_post_entry(trades[PE_entry_date])
            else:
                pass

        self.dataHist.save()

        return trades



class DataHist:
    """Class to access historical data"""

    def __init__(self, dataFile):
        """Populate all columns and bars"""

        self.fileName = dataFile
        self.wb = xl.load_workbook(self.fileName)

        # Read column headers

        self.wsc = self.wb.get_sheet_by_name('columns')  # columns tab

        self.header = {}  # format: {'column name': column number, ..... ]

        col = 1
        while self.wsc.cell(row=1, column=col).value is not None:
            self.header[self.wsc.cell(row=1, column=col).value] = col - 1
            col += 1

        # Populate all data

        self.wsd = self.wb.get_sheet_by_name('data') # data tab

        self.value = {}  # format: {'column name' : [row1_val, row2_val....], ......}

        for name, position in self.header.items():
            self.value[name] = self.xl_column(position)

        self.barindex = {}

        if 'Date' in self.header:
            for i in range(0, len(self.value['Date'])):
                self.barindex[self.value['Date'][i]] = i


        self.barcount = len(self.xl_column(0)) # number of bars in data tab


    def save(self):

        self.wb.save(self.fileName)


    def xl_column(self, col):
        """Return numpy for each column"""

        return np.array([r[col].value for r in self.wsd.iter_rows()])


    def bar(self, i):
        """ Return all columns in a single bar
        Return Format: {'column name' : value, .....}
        Call Format: bar(bar_index)
        """

        barvalues = {}

        for item in self.value:
            barvalues[item] = self.value[item][i]

        return barvalues

    def date_bar(self, date):
        """ Return all columns in a single bar
        Return Format: {'column name' : value, .....}
        Call Format: date_bar(date in YYYY-MM-DD)
        """
        return self.bar(self.barindex[date])

    def add_column_names(self, *columns):
        """Add new column names in columns tab"""

        next_column = len(self.header) + 1

        for column in columns:
            self.wsc.cell(row=1, column=next_column).value = column
            self.header[column] = next_column - 1
            next_column += 1


    def write_data_by_bar(self, bar, data):

        for column, value in data.items():
            self.wsd.cell(row=bar, column=self.header[column] + 1).value = value


    def write_data_by_date(self, date, data):

        for column, value in data.items():
            self.wsd.cell(row=self.barindex[date] + 1, column=self.header[column] + 1).value = value




